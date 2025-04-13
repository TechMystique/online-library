from django.views.generic import ListView, CreateView
from django.urls import reverse_lazy
from .models import Author, Book, BorrowRecord
from .forms import AuthorForm, BookForm, BorrowRecordForm
from django.http import HttpResponse
import openpyxl

from django.views.generic import TemplateView

class HomePageView(TemplateView):
    template_name = 'library/home.html'


# Create Views
class AuthorCreateView(CreateView):
    model = Author
    form_class = AuthorForm
    template_name = 'library/form.html'
    success_url = reverse_lazy('author-list')

class BookCreateView(CreateView):
    model = Book
    form_class = BookForm
    template_name = 'library/form.html'
    success_url = reverse_lazy('book-list')

class BorrowRecordCreateView(CreateView):
    model = BorrowRecord
    form_class = BorrowRecordForm
    template_name = 'library/form.html'
    success_url = reverse_lazy('borrow-list')

# List Views
class AuthorListView(ListView):
    model = Author
    paginate_by = 10
    template_name = 'library/author_list.html'

class BookListView(ListView):
    model = Book
    paginate_by = 10
    template_name = 'library/book_list.html'

class BorrowRecordListView(ListView):
    model = BorrowRecord
    paginate_by = 10
    template_name = 'library/borrow_list.html'

# Export View
def export_to_excel(request):
    wb = openpyxl.Workbook()
    
    # Export Authors
    ws1 = wb.active
    ws1.title = "Authors"
    ws1.append(['Name', 'Email', 'Bio'])
    for author in Author.objects.all():
        ws1.append([author.name, author.email, author.bio])

    # Export Books
    ws2 = wb.create_sheet(title="Books")
    ws2.append(['Title', 'Genre', 'Published Date', 'Author'])
    for book in Book.objects.all():
        ws2.append([book.title, book.genre, str(book.published_date), book.author.name])

    # Export Borrow Records
    ws3 = wb.create_sheet(title="Borrow Records")
    ws3.append(['User Name', 'Book', 'Borrow Date', 'Return Date'])
    for record in BorrowRecord.objects.all():
        ws3.append([record.user_name, record.book.title, str(record.borrow_date), str(record.return_date)])

    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=library_data.xlsx'
    return response
